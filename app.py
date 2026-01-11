from flask import Flask, render_template, request, jsonify, session, redirect, url_for, Response, send_file
import random
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
import os
import threading
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import arabic_reshaper
from bidi.algorithm import get_display

# 🛠️ הגדרת seed קבוע לתוצאות עקביות
RANDOM_SEED = 42
from flask_cors import CORS
import hashlib
import secrets
from datetime import datetime
import sqlite3
import re
import copy
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from pricing.scrapers import (
    get_prices_shufersal,
    get_prices_victory,
    get_prices_from_rami_levy
)

from algorithm import MenuOptimizer

FONT_DIR = os.path.join(os.path.dirname(__file__), "fonts")

pdfmetrics.registerFont(
    TTFont("DavidLibre", os.path.join(FONT_DIR, "DavidLibre-Regular.ttf"))
)

pdfmetrics.registerFont(
    TTFont("DavidLibre-Bold", os.path.join(FONT_DIR, "DavidLibre-Bold.ttf"))
)

# ייבוא ההגדרות החדשות
from config import SECRET_KEY, PORT, ADMIN_EMAIL
from database import get_db_connection, init_database


app = Flask(__name__)
app.secret_key = SECRET_KEY
CORS(app)


last_prices_update = None

# ============================
# 📦 מערכת ניהול משימות אסינכרונית
# ============================
import threading

price_update_tasks = {}
tasks_lock = threading.Lock()

def is_admin():
    if 'user_id' not in session:
        return False

    conn, db_type = get_db_connection()
    cursor = conn.cursor()
    
    placeholder = "%s" if db_type == 'postgresql' else "?"
    cursor.execute(f"SELECT email FROM users WHERE id = {placeholder}", (session['user_id'],))
    row = cursor.fetchone()
    conn.close()

    if not row:
        return False

    email = row['email'] if db_type == 'postgresql' else row[0]
    return email == ADMIN_EMAIL



def add_price_update_task(food_id, food_name):
    """הוספת משימת עדכון מחירים לרשימה"""
    task_id = f"price_update_{food_id}_{datetime.now().timestamp()}"
    with tasks_lock:
        price_update_tasks[task_id] = {
            'food_id': food_id,
            'food_name': food_name,
            'status': 'pending',
            'created_at': datetime.now().isoformat(),
            'message': 'ממתין להתחלת עדכון מחירים'
        }
    return task_id

def rtl(text):
    reshaped = arabic_reshaper.reshape(text)
    return get_display(reshaped)


def update_task_status(task_id, status, message=None):
    """עדכון סטטוס משימה"""
    with tasks_lock:
        if task_id in price_update_tasks:
            price_update_tasks[task_id]['status'] = status
            if message:
                price_update_tasks[task_id]['message'] = message
            if status in ['completed', 'failed']:
                price_update_tasks[task_id]['completed_at'] = datetime.now().isoformat()

def update_prices_for_single_food_async(food_id, task_id):
    """עדכון מחירים אסינכרוני לפריט בודד
    באותה לוגיקה של update_all_prices (רשימה בגודל 1)
    """
    try:
        # מציאת הפריט
        food = None
        for f in foods_db:
            if f['id'] == food_id:
                food = f
                break

        if not food:
            print(f"❌ [ASYNC] הפריט {food_id} לא נמצא")
            update_task_status(task_id, 'failed', 'הפריט לא נמצא')
            return

        product_name = food['name']
        separator = '=' * 60
        print(f"\n{separator}")
        print(f"🔍 [ASYNC] מתחיל חיפוש מחירים עבור: {product_name}")
        print(f"{separator}")

        update_task_status(task_id, 'running', f'מחפש מחירים עבור: {product_name}...')

        # === אותה לוגיקה של עדכון כללי – רשימה בגודל 1 ===
        product_names = [product_name]

        shufersal = get_prices_shufersal(product_names)
        victory   = get_prices_victory(product_names)
        rami      = get_prices_from_rami_levy(product_names)

        shuf = shufersal[0]
        vic  = victory[0]
        ram  = rami[0]

        # עדכון המחירים (בדיוק כמו update_all_prices)
        food['prices']['shufersal'] = shuf * 100 if shuf > 0 else None
        food['prices']['victory']   = vic  * 100 if vic  > 0 else None
        food['prices']['rami_levy'] = ram  * 100 if ram  > 0 else None

        print(f"\n📊 סיכום מחירים עבור {product_name}:")
        if shuf > 0:
            print(f"   • שופרסל: {shuf:.2f} ₪")
        else:
            print(f"   • שופרסל: לא נמצא")

        if vic > 0:
            print(f"   • ויקטורי: {vic:.2f} ₪")
        else:
            print(f"   • ויקטורי: לא נמצא")

        if ram > 0:
            print(f"   • רמי לוי: {ram:.2f} ₪")
        else:
            print(f"   • רמי לוי: לא נמצא")

        print(f"{separator}\n")

        update_task_status(task_id, 'completed', 'המחירים עודכנו בהצלחה')

    except Exception as e:
        print(f"\n❌ [ASYNC] שגיאה בעדכון מחירים: {e}")
        update_task_status(task_id, 'failed', f'שגיאה בעדכון מחירים: {str(e)}')


def start_price_update_task(food_id, food_name):
    """התחלת משימת עדכון מחירים ב-thread נפרד"""
    task_id = add_price_update_task(food_id, food_name)
    
    # יצירת thread חדש לעדכון המחירים
    thread = threading.Thread(
        target=update_prices_for_single_food_async,
        args=(food_id, task_id),
        daemon=True
    )
    thread.start()
    
    return task_id

def get_task_status(task_id):
    """קבלת סטטוס משימה"""
    with tasks_lock:
        return price_update_tasks.get(task_id)


def update_prices_by_names(product_names):
    """
    לוגיקת מחירים משותפת
    משמשת גם לעדכון כללי וגם לפריט בודד
    """
    shufersal = get_prices_shufersal(product_names)
    victory   = get_prices_victory(product_names)
    rami      = get_prices_from_rami_levy(product_names)

    return shufersal, victory, rami


def init_db():
    conn, db_type = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fullname TEXT NOT NULL,
            email TEXT UNIQUE NOT NULL,
            phone TEXT NOT NULL,
            password TEXT NOT NULL,
            created_at TEXT NOT NULL
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS foods (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            protein REAL,
            calories REAL,
            carbs REAL,
            fat REAL,
            category TEXT,
            allowed_meals TEXT,
            price_manual REAL,
            price_shufersal REAL,
            price_rami_levy REAL,
            price_victory REAL,
            active_price_source TEXT
        )
    ''')

    conn.commit()
    conn.close()


def get_default_foods():
    foods = [
        # ... כל הרשימה ששלחת ליך ...
    ]

    # חשוב מאוד!
    # אם אתה מעדכן עם prices כמו בקוד של ה-seed:
    for f in foods:
        f["prices"] = {
            "manual": f.pop("price", None),
            "shufersal": None,
            "rami_levy": None,
            "victory": None
        }
        f["active_price_source"] = "manual"

    return foods

def seed_foods_if_empty():
    conn, db_type = get_db_connection()
    cur = conn.cursor()

    # אם כבר יש מזונות בטבלה - לא עושים כלום
    cur.execute("SELECT COUNT(*) FROM foods")
    count = cur.fetchone()[0]
    if count > 0:
        conn.close()
        return

    # מכניסים את ברירת המחדל מהקוד
    foods = get_default_foods()

    for f in foods:
        allowed_meals_str = json.dumps(f.get("allowed_meals", []), ensure_ascii=False)

        prices = f.get("prices", {})
        price_manual = prices.get("manual", None)
        price_shufersal = prices.get("shufersal", None)
        price_rami_levy = prices.get("rami_levy", None)
        price_victory = prices.get("victory", None)

        active_src = f.get("active_price_source", "manual")

        cur.execute("""
            INSERT INTO foods (
                id, name, protein, calories, carbs, fat, category, allowed_meals,
                price_manual, price_shufersal, price_rami_levy, price_victory, active_price_source
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            f["id"], f["name"], f["protein"], f["calories"], f["carbs"], f["fat"], f["category"], allowed_meals_str,
            price_manual, price_shufersal, price_rami_levy, price_victory, active_src
        ))

    conn.commit()
    conn.close()

init_db()
seed_foods_if_empty()


# =========================
# מזונות ברירת מחדל
# =========================
def get_default_foods():
    foods = [
        {"id": "1", "name": "ביצים", "protein": 12.6, "calories": 155, "carbs": 1.0, "fat": 11.0, "price": 4.16, "category": "protein", "allowed_meals": ["breakfast", "dinner"]},
        {"id": "2", "name": "קוטג'", "protein": 11.0, "calories": 100, "carbs": 3.0, "fat": 5.0, "price": 3.50, "category": "protein", "allowed_meals": ["breakfast", "dinner"]},
        {"id": "3", "name": "יוגורט", "protein": 5.0, "calories": 70, "carbs": 6.0, "fat": 3.0, "price": 3.00, "category": "protein", "allowed_meals": ["breakfast", "snacks"]},
        {"id": "4", "name": "גבינה לבנה", "protein": 10.0, "calories": 120, "carbs": 2.0, "fat": 5.0, "price": 4.00, "category": "protein", "allowed_meals": ["breakfast"]},
        {"id": "5", "name": "חלב", "protein": 3.3, "calories": 64, "carbs": 5.0, "fat": 3.5, "price": 0.50, "category": "protein", "allowed_meals": ["breakfast"]},
        {"id": "6", "name": "גבינה צהובה", "protein": 25.0, "calories": 402, "carbs": 1.3, "fat": 33.0, "price": 5.00, "category": "protein", "allowed_meals": ["breakfast", "dinner"]},
        {"id": "7", "name": "גבינת שמנת", "protein": 6.0, "calories": 342, "carbs": 4.0, "fat": 34.0, "price": 4.50, "category": "protein", "allowed_meals": ["breakfast"]},
        {"id": "8", "name": "חזה עוף", "protein": 31.0, "calories": 165, "carbs": 0.0, "fat": 3.6, "price": 4.00, "category": "protein", "allowed_meals": ["lunch"]},
        {"id": "9", "name": "חזה הודו", "protein": 29.0, "calories": 135, "carbs": 0.0, "fat": 1.0, "price": 4.50, "category": "protein", "allowed_meals": ["lunch"]},
        {"id": "10", "name": "טונה בשמן", "protein": 26.0, "calories": 198, "carbs": 0.0, "fat": 8.0, "price": 2.50, "category": "protein", "allowed_meals": ["lunch", "dinner"]},
        {"id": "11", "name": "טונה במים", "protein": 28.0, "calories": 116, "carbs": 0.0, "fat": 0.8, "price": 2.80, "category": "protein", "allowed_meals": ["lunch", "dinner"]},
        {"id": "12", "name": "סלמון", "protein": 20.0, "calories": 208, "carbs": 0.0, "fat": 13.0, "price": 8.00, "category": "protein", "allowed_meals": ["lunch"]},
        {"id": "13", "name": "בשר בקר טחון", "protein": 26.0, "calories": 250, "carbs": 0.0, "fat": 15.0, "price": 5.00, "category": "protein", "allowed_meals": ["lunch"]},
        {"id": "14", "name": "סטייק בקר", "protein": 25.0, "calories": 271, "carbs": 0.0, "fat": 19.0, "price": 7.00, "category": "protein", "allowed_meals": ["lunch"]},
        {"id": "15", "name": "כבד עוף", "protein": 16.9, "calories": 119, "carbs": 0.9, "fat": 4.8, "price": 3.00, "category": "protein", "allowed_meals": ["lunch"]},
        {"id": "16", "name": "נקניק עוף", "protein": 12.0, "calories": 180, "carbs": 3.0, "fat": 15.0, "price": 3.50, "category": "protein", "allowed_meals": ["lunch"]},
        {"id": "17", "name": "נקניקיות הודו", "protein": 13.0, "calories": 170, "carbs": 2.0, "fat": 13.0, "price": 3.80, "category": "protein", "allowed_meals": ["lunch"]},
  
        # פחמימות
        {"id": "21", "name": "לחם מלא", "protein": 9.0, "calories": 247, "carbs": 41.0, "fat": 3.4, "price": 1.00, "category": "carb", "allowed_meals": ["breakfast", "dinner"]},
        {"id": "22", "name": "לחם לבן", "protein": 8.9, "calories": 265, "carbs": 49.0, "fat": 3.2, "price": 0.80, "category": "carb", "allowed_meals": ["breakfast", "dinner"]},
        {"id": "23", "name": "פיתה", "protein": 8.2, "calories": 275, "carbs": 55.7, "fat": 1.2, "price": 0.70, "category": "carb", "allowed_meals": ["breakfast", "dinner"]},
        {"id": "24", "name": "אורז לבן", "protein": 2.7, "calories": 130, "carbs": 28.0, "fat": 0.3, "price": 0.60, "category": "carb", "allowed_meals": ["lunch", "dinner"]},
        {"id": "25", "name": "אורז מלא", "protein": 2.6, "calories": 111, "carbs": 23.0, "fat": 0.9, "price": 0.80, "category": "carb", "allowed_meals": ["lunch", "dinner"]},
        {"id": "26", "name": "פסטה", "protein": 5.0, "calories": 131, "carbs": 25.0, "fat": 1.1, "price": 0.70, "category": "carb", "allowed_meals": ["lunch", "dinner"]},
        {"id": "27", "name": "קוסקוס", "protein": 3.8, "calories": 112, "carbs": 23.0, "fat": 0.2, "price": 0.65, "category": "carb", "allowed_meals": ["lunch", "dinner"]},
        {"id": "28", "name": "בורגול", "protein": 3.1, "calories": 83, "carbs": 18.6, "fat": 0.2, "price": 0.75, "category": "carb", "allowed_meals": ["lunch", "dinner"]},
        {"id": "29", "name": "תפוח אדמה", "protein": 2.0, "calories": 77, "carbs": 17.0, "fat": 0.1, "price": 0.40, "category": "carb", "allowed_meals": ["lunch", "dinner"]},
        {"id": "30", "name": "בטטה", "protein": 1.6, "calories": 86, "carbs": 20.0, "fat": 0.1, "price": 0.60, "category": "carb", "allowed_meals": ["lunch", "dinner"]},
        {"id": "31", "name": "קינואה", "protein": 4.4, "calories": 120, "carbs": 21.3, "fat": 1.9, "price": 1.50, "category": "carb", "allowed_meals": ["lunch", "dinner"]},
        {"id": "32", "name": "שיבולת שועל", "protein": 2.4, "calories": 71, "carbs": 12.0, "fat": 1.4, "price": 0.50, "category": "carb", "allowed_meals": ["breakfast"]},
        {"id": "33", "name": "קורנפלקס", "protein": 7.0, "calories": 357, "carbs": 84.0, "fat": 0.9, "price": 1.20, "category": "carb", "allowed_meals": ["breakfast"]},
        {"id": "34", "name": "גרנולה", "protein": 10.0, "calories": 471, "carbs": 64.0, "fat": 20.0, "price": 2.00, "category": "carb", "allowed_meals": ["breakfast", "snacks"]},
        
        # ירקות - שימו לב! הוספנו "breakfast" לגזר, חסה, ופלפל אדום!
        {"id": "35", "name": "מלפפון", "protein": 0.7, "calories": 10, "carbs": 4.0, "fat": 0.0, "price": 1.00, "category": "vegetable", "allowed_meals": ["breakfast", "lunch", "dinner"]},
        {"id": "36", "name": "עגבנייה", "protein": 0.9, "calories": 20, "carbs": 5.0, "fat": 0.0, "price": 1.00, "category": "vegetable", "allowed_meals": ["breakfast", "lunch", "dinner"]},
        {"id": "37", "name": "חסה", "protein": 1.4, "calories": 15, "carbs": 3.0, "fat": 0.0, "price": 0.80, "category": "vegetable", "allowed_meals": ["breakfast", "lunch", "dinner"]},
        {"id": "38", "name": "פלפל אדום", "protein": 1.0, "calories": 31, "carbs": 6.0, "fat": 0.3, "price": 1.50, "category": "vegetable", "allowed_meals": ["breakfast", "lunch", "dinner"]},
        {"id": "39", "name": "פלפל ירוק", "protein": 0.9, "calories": 20, "carbs": 5.0, "fat": 0.2, "price": 1.20, "category": "vegetable", "allowed_meals": ["lunch", "dinner"]},
        {"id": "40", "name": "גזר", "protein": 0.9, "calories": 41, "carbs": 10.0, "fat": 0.2, "price": 0.60, "category": "vegetable", "allowed_meals": ["breakfast", "lunch", "dinner"]},
        {"id": "41", "name": "כרוב", "protein": 1.3, "calories": 25, "carbs": 6.0, "fat": 0.1, "price": 0.50, "category": "vegetable", "allowed_meals": ["lunch", "dinner"]},
        {"id": "42", "name": "ברוקולי", "protein": 2.8, "calories": 34, "carbs": 7.0, "fat": 0.4, "price": 1.50, "category": "vegetable", "allowed_meals": ["lunch", "dinner"]},
        {"id": "43", "name": "כרובית", "protein": 1.9, "calories": 25, "carbs": 5.0, "fat": 0.3, "price": 1.20, "category": "vegetable", "allowed_meals": ["lunch", "dinner"]},
        {"id": "44", "name": "תירס", "protein": 3.4, "calories": 86, "carbs": 19.0, "fat": 1.4, "price": 0.80, "category": "vegetable", "allowed_meals": ["lunch", "dinner"]},
        {"id": "45", "name": "אפונה", "protein": 5.4, "calories": 81, "carbs": 14.0, "fat": 0.4, "price": 0.90, "category": "vegetable", "allowed_meals": ["lunch", "dinner"]},
        {"id": "46", "name": "שעועית ירוקה", "protein": 1.8, "calories": 31, "carbs": 7.0, "fat": 0.2, "price": 1.00, "category": "vegetable", "allowed_meals": ["lunch", "dinner"]},
        
        # פירות
        {"id": "47", "name": "תפוח", "protein": 0.3, "calories": 52, "carbs": 14.0, "fat": 0.2, "price": 0.80, "category": "fruit", "allowed_meals": ["breakfast", "snacks"]},
        {"id": "48", "name": "בננה", "protein": 1.1, "calories": 89, "carbs": 23.0, "fat": 0.3, "price": 0.60, "category": "fruit", "allowed_meals": ["breakfast", "snacks"]},
        {"id": "49", "name": "תפוז", "protein": 0.9, "calories": 47, "carbs": 12.0, "fat": 0.1, "price": 0.70, "category": "fruit", "allowed_meals": ["breakfast", "snacks"]},
        {"id": "50", "name": "אגס", "protein": 0.4, "calories": 57, "carbs": 15.0, "fat": 0.1, "price": 0.90, "category": "fruit", "allowed_meals": ["breakfast", "snacks"]},
        {"id": "51", "name": "אפרסק", "protein": 0.9, "calories": 39, "carbs": 10.0, "fat": 0.3, "price": 1.00, "category": "fruit", "allowed_meals": ["breakfast", "snacks"]},
        {"id": "52", "name": "אבטיח", "protein": 0.6, "calories": 30, "carbs": 8.0, "fat": 0.2, "price": 0.40, "category": "fruit", "allowed_meals": ["snacks"]},
        {"id": "53", "name": "מלון", "protein": 0.8, "calories": 34, "carbs": 8.0, "fat": 0.2, "price": 0.50, "category": "fruit", "allowed_meals": ["snacks"]},
        {"id": "54", "name": "ענבים", "protein": 0.7, "calories": 69, "carbs": 18.0, "fat": 0.2, "price": 1.50, "category": "fruit", "allowed_meals": ["snacks"]},
        {"id": "55", "name": "תות שדה", "protein": 0.7, "calories": 32, "carbs": 8.0, "fat": 0.3, "price": 2.00, "category": "fruit", "allowed_meals": ["breakfast", "snacks"]},
        {"id": "56", "name": "קיווי", "protein": 1.1, "calories": 61, "carbs": 15.0, "fat": 0.5, "price": 1.20, "category": "fruit", "allowed_meals": ["breakfast", "snacks"]},
        {"id": "57", "name": "חמאת בוטנים", "protein": 25.0, "calories": 588.0, "carbs": 20.0, "fat": 50.0, "price": 3.50, "category": "fat", "allowed_meals": ["snacks"]},
        {"id": "58", "name": "שקדים", "protein": 21.0, "calories": 579.0, "carbs": 22.0, "fat": 50.0, "price": 3.00, "category": "fat", "allowed_meals": ["snacks"]},
        {"id": "59", "name": "אגוזי מלך", "protein": 15.0, "calories": 654.0, "carbs": 14.0, "fat": 65.0, "price": 3.20, "category": "fat", "allowed_meals": ["snacks"]},
        {"id": "60", "name": "טחינה", "protein": 17.0, "calories": 595.0, "carbs": 21.0, "fat": 53.0, "price": 2.80, "category": "fat", "allowed_meals": ["snacks"]},
        {"id": "61", "name": "אבוקדו", "protein": 2.0, "calories": 160.0, "carbs": 9.0, "fat": 15.0, "price": 2.00, "category": "fat", "allowed_meals": ["snacks"]},
        {"id": "62", "name": "זרעי צ'יה", "protein": 17.0, "calories": 486.0, "carbs": 41.0, "fat": 31.0, "price": 1.80, "category": "fat", "allowed_meals": ["snacks"]},
        {"id": "63", "name": "זרעי פשתן", "protein": 18.0, "calories": 534.0, "carbs": 29.0, "fat": 42.0, "price": 1.60, "category": "fat", "allowed_meals": ["snacks"]},
        {"id": "64", "name": "יוגורט 0%", "protein": 10.0, "calories": 59.0, "carbs": 3.6, "fat": 0.4, "price": 2.50, "category": "fat", "allowed_meals": ["snacks"]},
        {"id": "65", "name": "יוגורט 3%", "protein": 5.0, "calories": 61.0, "carbs": 4.7, "fat": 3.0, "price": 2.50, "category": "fat", "allowed_meals": ["snacks"]},
        {"id": "66", "name": "תמר", "protein": 2.5, "calories": 277.0, "carbs": 75.0, "fat": 0.2, "price": 0.90, "category": "fruit", "allowed_meals": ["snacks"]}
    ]
  
  

    

    for f in foods:
        f["prices"] = {
            "manual": f.pop("price"),
            "shufersal": None,
            "rami_levy": None,
            "victory": None
        }
        f["active_price_source"] = "manual"

    return foods


foods_db = get_default_foods()

def update_all_prices():
    global last_prices_update

    product_names = [food["name"] for food in foods_db]

    shufersal = get_prices_shufersal(product_names)
    victory = get_prices_victory(product_names)
    rami = get_prices_from_rami_levy(product_names)

    for i, food in enumerate(foods_db):
        food["prices"]["shufersal"] = shufersal[i] * 100 if shufersal[i] > 0 else None
        food["prices"]["victory"] = victory[i] * 100 if victory[i] > 0 else None
        food["prices"]["rami_levy"] = rami[i] * 100 if rami[i] > 0 else None

    last_prices_update = datetime.now()
    print(f"🕒 מחירים עודכנו ב‑{last_prices_update}")







def run_optimizer_for_all_price_sources(foods_db, user_params, price_sources):
    store_totals = {}
    random.seed(42)

    best_result = None
    best_source = None
    best_cost = None

    for source in price_sources:
        foods_copy = []

        for f in foods_db:
            price = f["prices"].get(source)

            # אם אין מחיר – ממצעים מסופרים אחרים (בלי manual)
            if price is None:
                other_prices = [
                    p for k, p in f["prices"].items()
                    if k != "manual" and p is not None
                ]
                if not other_prices:
                    continue
                price = sum(other_prices) / len(other_prices)

            f_copy = copy.deepcopy(f)
            f_copy["prices"] = {"temp": price}
            f_copy["active_price_source"] = "temp"
            foods_copy.append(f_copy)

        if not foods_copy:
            continue

        optimizer = MenuOptimizer(foods_copy, user_params)
        optimizer.allowed_foods = copy.deepcopy(optimizer.original_allowed_foods)
        result = optimizer.generate_menu(num_days=user_params["num_days"])

        if result and result.get("success"):
            total_cost = result["total_cost"]
            store_totals[source] = total_cost

            print(f" חנות: {source}, מחיר כולל: {total_cost:.2f}")

            if best_cost is None or total_cost < best_cost:
                best_result = result
                best_source = source
                best_cost = total_cost

    if best_source:
        print(f"\n✅ החנות שנבחרה: {best_source} ({best_cost:.2f} ₪)")
        

    print("\n מחירים כוללים לפי סופר:")
    for store, cost in store_totals.items():
        print(f" {store}: {round(cost, 2)} ₪")

    return best_result, best_source





def nightly_price_update():
    print("🌙 התחל עדכון מחירים אוטומטי")
    update_all_prices()
    print("✅ עדכון מחירים לילי הסתיים")




@app.route("/api/prices/update", methods=["POST"])
def update_prices():
    if not is_logged_in():
        return jsonify({"success": False}), 401

    try:
        update_all_prices()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# =========================
# עזר
# =========================


def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()


def is_logged_in():
    return 'user_id' in session


# =========================
# נתיבי עמודים
# =========================

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/login')
def login():
    return render_template('login.html')


@app.route('/login', methods=['POST'])
def login_post():
    data = request.get_json()
    email = data.get('email', '').strip()
    password = data.get('password', '')

    if not email or not password:
        return jsonify({'success': False, 'message': 'נא למלא את כל השדות'})

    conn, db_type = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT id, password FROM users WHERE email = ?', (email,))
    user = cursor.fetchone()
    conn.close()

    if user and user[1] == hash_password(password):
        session['user_id'] = user[0]
        return jsonify({'success': True})

    else:
        return jsonify({'success': False, 'message': 'כתובת אימייל לא נכונה או סיסמה שגויה'})


@app.route("/admin/users")
def admin_users():
    if not is_admin():
        return "Access denied", 403

    conn, db_type = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id, fullname, email
        FROM users
        ORDER BY id DESC
    """)
    users = cursor.fetchall()

    cursor.execute("SELECT COUNT(*) FROM users")
    total_users = cursor.fetchone()[0]

    conn.close()

    return render_template(
        "admin_users.html",
        users=users,
        total_users=total_users
    )


@app.route('/register')
def register():
    return render_template('register.html')


@app.route('/register', methods=['POST'])
def register_post():
    data = request.get_json()
    fullname = data.get('fullname', '').strip()
    email = data.get('email', '').strip()
    phone = data.get('phone', '').strip()
    password = data.get('password', '')

    # ולידציה
    if not fullname or not email or not phone or not password:
        return jsonify({'success': False, 'message': 'נא למלא את כל השדות'})

    # ולידצית אימייל
    email_regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    if not re.match(email_regex, email):
        return jsonify({'success': False, 'message': 'כתובת האימייל לא תקינה'})

    # ולידצית טלפון
    phone_regex = r'^0[0-9]{8,9}$'
    if not re.match(phone_regex, phone):
        return jsonify({'success': False, 'message': 'מספר טלפון לא תקין'})

    # ולידצית סיסמה (לפחות 6 תווים)
    if len(password) < 6:
        return jsonify({'success': False, 'message': 'הסיסמה חייבת להכיל לפחות 6 תווים'})

    conn, db_type = get_db_connection()
    cursor = conn.cursor()

    # בדיקה אם האימייל כבר קיים
    cursor.execute('SELECT id FROM users WHERE email = ?', (email,))
    if cursor.fetchone():
        conn.close()
        return jsonify({'success': False, 'message': 'כתובת האימייל כבר קיימת במערכת'})

    # יצירת משתמש חדש
    try:
        cursor.execute('''
            INSERT INTO users (fullname, email, phone, password, created_at)
            VALUES (?, ?, ?, ?, ?)
        ''', (fullname, email, phone, hash_password(password), datetime.now().isoformat()))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'ההרשמה בוצעה בהצלחה'})
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'message': 'שגיאה בהרשמה, אנא נסה שוב'})


@app.route('/dashboard')
def dashboard():
    if not is_logged_in():
        return redirect(url_for('login'))
    return render_template('dashboard.html')


@app.route('/food-management')
def food_management():
    if not is_logged_in():
        return redirect(url_for('login'))
    return render_template('food-management.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))


# =========================
# API בדיקת אימות
# =========================

@app.route('/check-auth')
def check_auth():
    if is_logged_in():
        return jsonify({'authenticated': True})
    return jsonify({'authenticated': False}), 401


# =========================
# API מזון
# =========================
@app.route('/api/foods', methods=['GET'])
def get_foods():
    if not is_logged_in():
        return jsonify({'success': False}), 401
    return jsonify({'success': True, 'data': foods_db})


@app.route('/api/foods', methods=['POST'])
def add_food():
    if not is_logged_in():
        return jsonify({'success': False}), 401

    data = request.get_json()
    global foods_db

    existing_ids = [
        int(f["id"]) for f in foods_db
        if str(f["id"]).isdigit()
    ]

    new_id = str(max(existing_ids) + 1 if existing_ids else 1)


    new_food = {
        'id': new_id,
        'name': data['name'],
        'protein': float(data['protein']),
        'calories': float(data['calories']),
        'carbs': float(data['carbs']),
        'fat': float(data['fat']),
        'prices': {
            'manual': float(data['price']),
            'shufersal': None,
            'rami_levy': None,
            'victory': None
        },
        'active_price_source': 'manual',
        'category': data['category'],
        'allowed_meals': data['allowed_meals']
    }

    foods_db.append(new_food)

    # 🔍 התחלת חיפוש מחירים אסינכרוני למזון החדש
    task_id = start_price_update_task(new_id, new_food['name'])

    return jsonify({
        'success': True,
        'message': 'המזון נשמר בהצלחה',
        'data': new_food,
        'task_id': task_id
    })


    
@app.route("/admin/export-users")
def export_users():
    if not is_admin():
        return "Access denied", 403

    conn, db_type = get_db_connection()
    cursor = conn.cursor()

    # ❌ בלי created_at
    cursor.execute("""
        SELECT id, fullname, email
        FROM users
        ORDER BY id
    """)
    users = cursor.fetchall()
    conn.close()

    # יצירת קובץ אקסל
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"

    # כותרות – בלי תאריך
    headers = ["ID", "שם מלא", "אימייל"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    # נתונים
    for u in users:
        ws.append(list(u))

    # שמירה לזיכרון
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="users.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )



@app.route('/api/foods/<food_id>', methods=['PUT'])
def update_food(food_id):
    if not is_logged_in():
        return jsonify({'success': False}), 401

    data = request.get_json()
    global foods_db

    # 1️⃣ מציאת הפריט והאינדקס שלו
    old_index = None
    for idx, food in enumerate(foods_db):
        if food['id'] == food_id:
            old_index = idx
            break

    if old_index is None:
        return jsonify({'success': False}), 404

    # 2️⃣ מחיקת הפריט הישן
    foods_db.pop(old_index)

    # 3️⃣ יצירת ID חדש (max + 1)
    existing_ids = [
        int(f["id"]) for f in foods_db
        if str(f["id"]).isdigit()
    ]
    new_id = str(max(existing_ids) + 1 if existing_ids else 1)

    # 4️⃣ יצירת פריט חדש מהנתונים שנשלחו
    new_food = {
        'id': new_id,
        'name': data['name'],
        'protein': float(data['protein']),
        'calories': float(data['calories']),
        'carbs': float(data['carbs']),
        'fat': float(data['fat']),
        'prices': {
            'manual': float(data['price']),
            'shufersal': None,
            'rami_levy': None,
            'victory': None
        },
        'active_price_source': 'manual',
        'category': data['category'],
        'allowed_meals': data['allowed_meals']
    }

    # 5️⃣ החזרה בדיוק לאותו מקום ברשימה
    foods_db.insert(old_index, new_food)

    # 6️⃣ התחלת עדכון מחירים אסינכרוני (כמו בהוספה)
    task_id = start_price_update_task(new_id, new_food['name'])

    # 7️⃣ החזרת תשובה ל־Frontend
    return jsonify({
        'success': True,
        'message': 'המזון נשמר בהצלחה',
        'data': new_food,
        'task_id': task_id
    })

@app.route('/export-shopping-list/excel', methods=['POST'])
def export_shopping_list_excel():
    if not is_logged_in():
        return jsonify({'error': 'Not authenticated'}), 401

    data = request.json
    shopping = {}

    for day in data['days']:
        for meal in ['breakfast', 'lunch', 'dinner', 'snacks']:
            for item in day.get(meal, []):
                shopping[item['name']] = shopping.get(item['name'], 0) + item['amount']

    wb = Workbook()
    ws = wb.active
    ws.title = "רשימת קניות"

    ws.append(["מוצר", "כמות (גרם)"])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for name, grams in shopping.items():
        ws.append([name, round(grams)])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="shopping_list.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route('/api/foods/<food_id>', methods=['DELETE'])
def delete_food(food_id):
    if not is_logged_in():
        return jsonify({'success': False}), 401

    global foods_db
    foods_db = [food for food in foods_db if food['id'] != food_id]
    return jsonify({'success': True})


@app.route('/export-foods')
def export_foods():
    if not is_logged_in():
        return redirect(url_for('login'))

    wb = Workbook()
    ws = wb.active
    ws.title = "מזונות"

    headers = ['ID', 'שם', 'חלבון', 'קלוריות', 'פחמימות', 'שומנים', 'מחיר פעיל', 'מקור מחיר']
    ws.append(headers)

    for food in foods_db:
        ws.append([
            food['id'],
            food['name'],
            food['protein'],
            food['calories'],
            food['carbs'],
            food['fat'],
            food['prices'][food['active_price_source']],
            food['active_price_source']
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment;filename=nutrioptimal_foods.xlsx'}
    )


@app.route('/api/prices/last-update')
def get_last_price_update():
    if not is_logged_in():
        return jsonify({'success': False}), 401

    return jsonify({
        'success': True,
        'last_update': last_prices_update.isoformat() if last_prices_update else None
    })



# ============================
# 🔍 API לבדיקת סטטוס משימות עדכון מחירים
# ============================
@app.route('/api/prices/task/<task_id>', methods=['GET'])
def get_price_update_task_status(task_id):
    if not is_logged_in():
        return jsonify({'success': False}), 401
    
    task = get_task_status(task_id)
    if not task:
        return jsonify({'success': False, 'error': 'Task not found'}), 404
    
    return jsonify({
        'success': True,
        'task': task
    })

@app.route('/export-menu', methods=['POST'])
def export_menu():
    if not is_logged_in():
        return jsonify({'error': 'Not authenticated'}), 401

    try:
        data = request.json
        days = data.get('days', [])

        wb = Workbook()
        wb.remove(wb.active)

        # 🎨 צבעים לארוחות
        meal_colors = {
            'בוקר': 'FFF2CC',
            'צהריים': 'DDEBF7',
            'ערב': 'E2EFDA',
            'חטיפים': 'FCE4D6'
        }

        title_font = Font(size=16, bold=True)
        meal_font = Font(size=13, bold=True)
        header_font = Font(bold=True)
        center = Alignment(horizontal='center', vertical='center')

        summary_ws = wb.create_sheet("סיכום יומי")
        summary_ws.append(['יום', 'קלוריות', 'חלבון', 'עלות יומית (₪)'])

        for day_idx, day in enumerate(days, 1):
            ws = wb.create_sheet(f"יום {day_idx}")

            # כותרת יום
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
            ws['A1'] = f"תפריט ליום {day_idx}"
            ws['A1'].font = title_font
            ws['A1'].alignment = center

            row = 3

            meals_map = {
                'breakfast': 'בוקר',
                'lunch': 'צהריים',
                'dinner': 'ערב',
                'snacks': 'חטיפים'
            }

            daily_cost = 0

            for meal_key, meal_name in meals_map.items():
                items = day.get(meal_key, [])
                if not items:
                    continue

                # כותרת ארוחה
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
                ws[f'A{row}'] = meal_name
                ws[f'A{row}'].font = meal_font
                ws[f'A{row}'].alignment = center
                ws[f'A{row}'].fill = PatternFill(
                    'solid', fgColor=meal_colors.get(meal_name)
                )
                row += 1

                # כותרות טבלה
                ws.append(['מזון', 'כמות (גרם)', 'קלוריות', 'מחיר (₪)'])
                for c in ws[row]:
                    c.font = header_font
                    c.alignment = center
                row += 1

                for item in items:
                    food = next(
                        (f for f in foods_db if f['name'] == item['name']),
                        None
                    )

                    calories = (
                        food['calories'] * item['amount'] / 100
                        if food else 0
                    )

                    price = (
                        food['prices'][food['active_price_source']] * item['amount'] / 100
                        if food else 0
                    )

                    daily_cost += price

                    ws.append([
                        item['name'],
                        round(item['amount'], 1),
                        round(calories, 1),
                        round(price, 2)
                    ])

                    for c in ws[row]:
                        c.alignment = center
                        c.fill = PatternFill(
                            'solid', fgColor=meal_colors.get(meal_name)
                        )
                    row += 1

                row += 1

            # ➕ שורה ריקה
            row += 1

            # 📊 סיכום יומי בתוך ה־Sheet
            ws.append(['סה״כ קלוריות', round(day['calories'], 1)])
            ws.append(['סה״כ חלבון', round(day['protein'], 1)])
            ws.append(['עלות יומית (₪)', round(daily_cost, 2)])

            # 🧮 סיכום מרכזי
            summary_ws.append([
                f"יום {day_idx}",
                round(day['calories'], 1),
                round(day['protein'], 1),
                round(daily_cost, 2)
            ])

            # התאמת רוחב עמודות
            for col in ['A', 'B', 'C', 'D']:
                ws.column_dimensions[col].width = 22

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name='menu_advanced.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    
    

    
@app.route('/export-menu/pdf', methods=['POST'])
def export_menu_pdf():
    if not is_logged_in():
        return jsonify({'error': 'Not authenticated'}), 401

    data = request.json
    days = data.get('days', [])

    buffer = BytesIO()

    # A4 לרוחב
    page_width, page_height = A4[1], A4[0]
    c = canvas.Canvas(buffer, pagesize=(page_width, page_height))

    margin_x = 20
    margin_y = 35

    days_per_page = 7
    card_gap = 8

    card_width = (
        page_width - 2 * margin_x - (days_per_page - 1) * card_gap
    ) / days_per_page
    card_height = page_height - 2 * margin_y

    meals_order = [
        ('breakfast', 'בוקר'),
        ('lunch', 'צהריים'),
        ('dinner', 'ערב'),
        ('snacks', 'תוספות')
    ]

    for page_start in range(0, len(days), days_per_page):
        page_days = days[page_start:page_start + days_per_page]

        for i, day in enumerate(page_days):
            # RTL – יום 1 מימין
            x = page_width - margin_x - (i + 1) * card_width - i * card_gap
            y_top = page_height - margin_y
            y_bottom = y_top - card_height

            # מסגרת כרטיס
            c.setStrokeColorRGB(1, 0.6, 0)
            c.roundRect(x, y_bottom, card_width, card_height, 12)

            # כותרת יום
            c.setFillColorRGB(1, 0.6, 0)
            c.roundRect(x, y_top - 32, card_width, 32, 12, fill=1)

            c.setFillColorRGB(0, 0, 0)
            c.setFont("DavidLibre-Bold", 12)
            c.drawCentredString(
                x + card_width / 2,
                y_top - 22,
                rtl(f"יום {page_start + i + 1}")
            )

            current_y = y_top - 48  # ⬅ רווח ברור אחרי כותרת יום

            for meal_key, meal_name in meals_order:
                items = day.get(meal_key, [])
                if not items:
                    continue

                # אם אין מקום – לא לצייר
                if current_y - 40 < y_bottom:
                    break

                # כותרת ארוחה
                c.setFillColorRGB(1, 0.88, 0.65)
                c.roundRect(
                    x + 6,
                    current_y - 22,
                    card_width - 12,
                    22,
                    7,
                    fill=1
                )

                c.setFillColorRGB(0, 0, 0)
                c.setFont("DavidLibre-Bold", 10)
                c.drawCentredString(
                    x + card_width / 2,
                    current_y - 15,
                    rtl(meal_name)
                )

                current_y -= 37  # ⬅ רווח חובה אחרי כותרת ארוחה

                # פריטי מזון
                c.setFont("DavidLibre-Bold", 9)

                for item in items:
                    if current_y - 12 < y_bottom:
                        break

                    line = f"{item['name']} – {round(item['amount'])} גרם"
                    c.drawRightString(
                        x + card_width - 10,
                        current_y,
                        rtl(line)
                    )
                    current_y -= 14  # ⬅ ריווח שורות ברור

                current_y -= 10  # ⬅ רווח בין ארוחות

        c.showPage()

    c.save()
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="menu_cards.pdf",
        mimetype="application/pdf"
    )






@app.route('/export-shopping-list/pdf', methods=['POST'])
def export_shopping_list_pdf():
    if not is_logged_in():
        return jsonify({'error': 'Not authenticated'}), 401

    data = request.json
    shopping = {}

    # איסוף רשימת קניות
    for day in data['days']:
        for meal in ['breakfast', 'lunch', 'dinner', 'snacks']:
            for item in day.get(meal, []):
                shopping[item['name']] = shopping.get(item['name'], 0) + item['amount']

    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    y = height - 50
    c.setFont("DavidLibre", 14)

    # כותרת בעברית (RTL)
    c.drawString(50, y, rtl("רשימת קניות"))
    y -= 30

    # שורות הרשימה
    for name, grams in shopping.items():
        line = f"{name} - {round(grams)} גרם"
        c.drawString(50, y, rtl(line))
        y -= 20

        # מעבר עמוד אם נגמר המקום
        if y < 50:
            c.showPage()
            c.setFont("DavidLibre", 14)
            y = height - 50
            c.drawString(50, y, rtl("רשימת קניות"))
            y -= 30

    c.save()
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="shopping_list.pdf",
        mimetype="application/pdf"
    )




@app.route('/calculate', methods=['POST'])
def calculate_menu():
    if not is_logged_in():
        return jsonify({'success': False}), 401

    try:
        data = request.get_json()
        selected_sources = data.get("price_sources", [])
        if not selected_sources:
            return jsonify({"success": False, "message": "לא נבחר מקור מחיר"}), 400



        
        user_params = {
            'num_days': int(data.get('num_days', 7)),
            'min_protein': float(data.get('min_protein', 56)),
            'max_protein': float(data.get('max_protein', 100)),
            'min_calories': float(data.get('min_calories', 1500)),
            'max_calories': float(data.get('max_calories', 2700)),
            'min_carbs': float(data.get('min_carbs', 150)),
            'max_carbs': float(data.get('max_carbs', 300)),
            'min_fat': float(data.get('min_fat', 50)),
            'max_fat': float(data.get('max_fat', 90))
        }

        result, chosen_source = run_optimizer_for_all_price_sources(
            foods_db,
            user_params,
            selected_sources
        )

        if not result:
            return jsonify({
                "success": False,
                "message": "לא נמצא פתרון לפי מקור מחיר זה"
            })


        if result['success']:
            # עיבוד התוצאות לפורמט שה-JavaScript מצפה לו
            processed_days = []
            total_protein = 0
            total_calories = 0
            total_carbs = 0
            total_fat = 0

            for day in result['days']:
                day_data = {
                    'breakfast': [],
                    'lunch': [],
                    'dinner': [],
                    'snacks': [],
                    'protein': day['nutrition']['protein'],
                    'calories': day['nutrition']['calories'],
                    'carbs': day['nutrition']['carbs'],
                    'fat': day['nutrition']['fat']
                }

                for meal_name in ['breakfast', 'lunch', 'dinner', 'snacks']:
                    for food_name, qty in day[meal_name]:
                        day_data[meal_name].append({
                            'name': food_name,
                            'amount': round(qty, 1)
                        })

                processed_days.append(day_data)
                total_protein += day['nutrition']['protein']
                total_calories += day['nutrition']['calories']
                total_carbs += day['nutrition']['carbs']
                total_fat += day['nutrition']['fat']

            avg_protein = total_protein / len(result['days'])
            avg_calories = total_calories / len(result['days'])
            avg_carbs = total_carbs / len(result['days'])
            avg_fat = total_fat / len(result['days'])
            
            print("📦 RESULT DATA (raw optimizer result):")
            print(json.dumps(result, indent=2, ensure_ascii=False))

            return jsonify({
                'success': True,
                'data': {
                    'days': processed_days,
                    'total_cost': result['total_cost'],
                    'avg_cost_per_day': result['avg_cost_per_day'],
                    'chosen_price_source': chosen_source,
                    'avg_protein': avg_protein,
                    'avg_calories': avg_calories,
                    'avg_carbs': avg_carbs,
                    'avg_fat': avg_fat
                    
                }
            })
        else:
            return jsonify({
                'success': False,
                'message': result.get('message', 'שגיאה בחישוב התפריט')
            })

    except Exception as e:
        print(f"Error in calculate_menu: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'שגיאה בחישוב התפריט: {str(e)}'
        }), 500


if os.environ.get("WERKZEUG_RUN_MAIN") == "true":
    scheduler = BackgroundScheduler()

    scheduler.add_job(
        func=nightly_price_update,
        trigger=CronTrigger(hour=3, minute=0),  # כל יום ב‑03:00
        id='nightly_price_update',
        replace_existing=True
    )

    scheduler.start()


if __name__ == '__main__':
    from config import DEBUG, PORT
    app.run(debug=DEBUG, port=PORT, host='0.0.0.0', threaded=True)